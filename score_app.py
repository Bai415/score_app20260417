# -*- coding: utf-8 -*-
"""
Created on Fri Apr 17 10:07:32 2026

@author: Lenovo
"""

import streamlit as st
import openpyxl as px1
from openpyxl.styles import PatternFill, colors, Font
import pandas as pd
from io import BytesIO

# ==================== 页面配置 ====================
st.set_page_config(
    page_title="试卷自动判分系统",
    page_icon="📊",
    layout="wide"
)

st.title("📊 试卷自动判分系统")
st.markdown("上传标准答案和学生答卷，系统自动判分并生成分析报告")

# ==================== 侧边栏：参数设置 ====================
with st.sidebar:
    st.header("⚙️ 评分参数设置")
    
    # 单选题参数
    st.subheader("单选题")
    single_points = st.number_input("每题分值", value=0.5, step=0.1, key="single_points")
    single_rows = st.text_input("行范围（格式：起始-结束）", value="1-21", key="single_rows")
    
    # 多选题参数
    st.subheader("多选题")
    multi_points = st.number_input("每题分值", value=0.7, step=0.1, key="multi_points")
    multi_rows = st.text_input("行范围（格式：起始-结束）", value="22-32", key="multi_rows")
    
    # 判断题参数
    st.subheader("判断题")
    judge_points = st.number_input("每题分值", value=0.3, step=0.1, key="judge_points")
    judge_rows = st.text_input("行范围（格式：起始-结束）", value="33-43", key="judge_rows")
    
    st.divider()
    st.caption("提示：行范围需与Excel中的实际布局一致")

# ==================== 主界面：文件上传 ====================
col1, col2 = st.columns(2)

with col1:
    st.subheader("📄 标准答案")
    std_file = st.file_uploader(
        "上传标准答案文件（.xlsx）",
        type=["xlsx"],
        key="std_file"
    )

with col2:
    st.subheader("👨‍🎓 学生答卷")
    student_files = st.file_uploader(
        "上传学生答卷文件（可多选）",
        type=["xlsx"],
        accept_multiple_files=True,
        key="student_files"
    )

# ==================== 判分核心函数 ====================
def parse_row_range(range_str):
    """解析行范围字符串，如'1-21'返回(1,21)"""
    start, end = range_str.split('-')
    return int(start), int(end)

def check_student(ws_std, ws_stu, single_range, multi_range, judge_range, 
                  single_pts, multi_pts, judge_pts):
    """
    对比学生答案与标准答案，返回错题列表和分数
    """
    single_wrong = []
    multi_wrong = []
    judge_wrong = []
    
    max_row = max(ws_std.max_row, ws_stu.max_row)
    max_col = max(ws_std.max_column, ws_stu.max_column)
    
    # 解析行范围
    s_start, s_end = parse_row_range(single_range)
    m_start, m_end = parse_row_range(multi_range)
    j_start, j_end = parse_row_range(judge_range)
    
    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            cell_std = ws_std.cell(i, j)
            cell_stu = ws_stu.cell(i, j)
            
            if cell_stu.value != cell_std.value:
                # 标记错误（仅当需要保存时才标记，Web版可选）
                
                # 根据行号判断题型
                if s_start <= i <= s_end:          # 单选题区域
                    title_cell = ws_std.cell(i-1, j)
                    try:
                        q_num = int(title_cell.value)
                        if 1 <= q_num <= 100:
                            single_wrong.append(q_num)
                    except:
                        pass
                elif m_start <= i <= m_end:        # 多选题区域
                    title_cell = ws_std.cell(i-1, j)
                    try:
                        q_num = int(title_cell.value)
                        if 1 <= q_num <= 50:
                            multi_wrong.append(q_num)
                    except:
                        pass
                elif j_start <= i <= j_end:        # 判断题区域
                    title_cell = ws_std.cell(i-1, j)
                    try:
                        q_num = int(title_cell.value)
                        if 1 <= q_num <= 50:
                            judge_wrong.append(q_num)
                    except:
                        pass
    
    # 计算总分
    score = 100 - len(single_wrong)*single_pts - len(multi_wrong)*multi_pts - len(judge_wrong)*judge_pts
    return single_wrong, multi_wrong, judge_wrong, score

# ==================== 执行判分 ====================
if std_file and student_files:
    if st.button("🚀 开始判分", type="primary"):
        with st.spinner("判分进行中，请稍候..."):
            try:
                # 加载标准答案
                wb_std = px1.load_workbook(BytesIO(std_file.read()))
                ws_std = wb_std['Sheet1']
                
                # 存储所有学生结果
                results = []
                all_single_wrong = []  # 用于统计群体错题
                all_multi_wrong = []
                all_judge_wrong = []
                
                # 获取参数
                single_pts = st.session_state.single_points
                multi_pts = st.session_state.multi_points
                judge_pts = st.session_state.judge_points
                single_range = st.session_state.single_rows
                multi_range = st.session_state.multi_rows
                judge_range = st.session_state.judge_rows
                
                # 逐个处理学生文件
                for stu_file in student_files:
                    wb_stu = px1.load_workbook(BytesIO(stu_file.read()))
                    ws_stu = wb_stu['Sheet1']
                    
                    single_wrong, multi_wrong, judge_wrong, score = check_student(
                        ws_std, ws_stu, single_range, multi_range, judge_range,
                        single_pts, multi_pts, judge_pts
                    )
                    
                    # 提取文件名（不含扩展名）
                    name = stu_file.name.replace('.xlsx', '')
                    
                    results.append({
                        "姓名": name,
                        "总分": round(score, 1),
                        "单选错误数": len(single_wrong),
                        "单选错题号": single_wrong,
                        "多选错误数": len(multi_wrong),
                        "多选错题号": multi_wrong,
                        "判断错误数": len(judge_wrong),
                        "判断错题号": judge_wrong
                    })
                    
                    all_single_wrong.extend(single_wrong)
                    all_multi_wrong.extend(multi_wrong)
                    all_judge_wrong.extend(judge_wrong)
                
                # ==================== 显示结果 ====================
                st.success(f"✅ 判分完成！共处理 {len(results)} 名学生")
                
                # 1. 成绩汇总表格
                st.subheader("📋 成绩汇总表")
                df_results = pd.DataFrame([
                    {k: v for k, v in r.items() if k not in ['单选错题号', '多选错题号', '判断错题号']}
                    for r in results
                ])
                st.dataframe(df_results, use_container_width=True)
                
                # 2. 详细错题报告
                st.subheader("📝 详细错题报告")
                for r in results:
                    with st.expander(f"{r['姓名']} - 总分：{r['总分']}分"):
                        col_a, col_b, col_c = st.columns(3)
                        with col_a:
                            st.metric("单选题错误", r['单选错误数'])
                            if r['单选错题号']:
                                st.write(f"错题号：{', '.join(map(str, sorted(r['单选错题号'])))}")
                        with col_b:
                            st.metric("多选题错误", r['多选错误数'])
                            if r['多选错题号']:
                                st.write(f"错题号：{', '.join(map(str, sorted(r['多选错题号'])))}")
                        with col_c:
                            st.metric("判断题错误", r['判断错误数'])
                            if r['判断错题号']:
                                st.write(f"错题号：{', '.join(map(str, sorted(r['判断错题号'])))}")
                
                # 3. 群体错题分析（错误率超过50%的题目）
                if len(results) > 1:
                    st.subheader("📊 群体错题分析")
                    threshold = len(results) / 2
                    
                    # 统计每道题的错误次数
                    from collections import Counter
                    single_counter = Counter(all_single_wrong)
                    multi_counter = Counter(all_multi_wrong)
                    judge_counter = Counter(all_judge_wrong)
                    
                    single_over = [q for q, cnt in single_counter.items() if cnt > threshold]
                    multi_over = [q for q, cnt in multi_counter.items() if cnt > threshold]
                    judge_over = [q for q, cnt in judge_counter.items() if cnt > threshold]
                    
                    if single_over or multi_over or judge_over:
                        st.write("以下题目错误率超过50%，建议重点讲解：")
                        if single_over:
                            st.write(f"🔴 单选题：{', '.join(map(str, sorted(single_over)))}")
                        if multi_over:
                            st.write(f"🟠 多选题：{', '.join(map(str, sorted(multi_over)))}")
                        if judge_over:
                            st.write(f"🟡 判断题：{', '.join(map(str, sorted(judge_over)))}")
                    else:
                        st.write("✅ 所有题目错误率均未超过50%")
                
                # 4. 下载报告（可选）
                st.subheader("💾 下载报告")
                report_text = ""
                for idx, r in enumerate(results, 1):
                    report_text += f"{idx}、{r['姓名']}总分：{r['总分']}分，"
                    report_text += f"单选错误个数：{r['单选错误数']}"
                    if r['单选错题号']:
                        report_text += f"（题号：{', '.join(map(str, sorted(r['单选错题号'])))})"
                    report_text += "；"
                    report_text += f"多选错误个数：{r['多选错误数']}"
                    if r['多选错题号']:
                        report_text += f"（题号：{', '.join(map(str, sorted(r['多选错题号'])))})"
                    report_text += "；"
                    report_text += f"判断错误个数：{r['判断错误数']}"
                    if r['判断错题号']:
                        report_text += f"（题号：{', '.join(map(str, sorted(r['判断错题号'])))})"
                    report_text += "。\n"
                
                st.download_button(
                    label="📥 下载成绩报告（TXT格式）",
                    data=report_text,
                    file_name="成绩报告.txt",
                    mime="text/plain"
                )
                
            except Exception as e:
                st.error(f"❌ 处理过程中出错：{str(e)}")
                st.info("请检查Excel文件格式是否正确（需要包含'Sheet1'工作表）")

else:
    st.info("👈 请先在左侧上传标准答案和学生答卷文件")

# ==================== 页脚 ====================
st.divider()
st.caption("试卷自动判分系统 | 基于Streamlit构建")